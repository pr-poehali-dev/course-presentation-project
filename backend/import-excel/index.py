import json
import os
import base64
import io
import psycopg2
from openpyxl import load_workbook


CATEGORY_MAP = {
    'ткани': 'fabrics',
    'ткань': 'fabrics',
    'портьерные': 'curtains',
    'портьерные ткани': 'curtains',
    'портьеры': 'curtains',
    'шторы': 'curtains',
    'тюлевые': 'tulle',
    'тюлевые ткани': 'tulle',
    'тюль': 'tulle',
    'мебельные': 'furniture',
    'мебельные ткани': 'furniture',
    'мебель': 'furniture',
    'домашний текстиль': 'home',
    'текстиль': 'home',
    'дом': 'home',
    'фурнитура': 'accessories',
    'аксессуары': 'accessories',
}


def normalize_category(value: str) -> str:
    if not value:
        return 'curtains'
    key = str(value).strip().lower()
    return CATEGORY_MAP.get(key, 'curtains')


def handler(event: dict, context) -> dict:
    """Импорт товаров из Excel-файла в каталог"""
    method = event.get('httpMethod', 'GET')

    cors_headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type',
        'Access-Control-Max-Age': '86400'
    }

    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': cors_headers, 'body': '', 'isBase64Encoded': False}

    if method != 'POST':
        return {'statusCode': 405, 'headers': {**cors_headers, 'Content-Type': 'application/json'}, 'body': json.dumps({'error': 'Method not allowed'}), 'isBase64Encoded': False}

    raw_body = event.get('body') or '{}'
    try:
        body = json.loads(raw_body)
    except Exception:
        body = {}
    file_data = body.get('file_data', '')

    if not file_data:
        return {'statusCode': 400, 'headers': {**cors_headers, 'Content-Type': 'application/json'}, 'body': json.dumps({'error': 'Файл не передан'}), 'isBase64Encoded': False}

    if ',' in file_data:
        file_data = file_data.split(',')[1]

    file_data = ''.join(file_data.split())
    padding = len(file_data) % 4
    if padding:
        file_data += '=' * (4 - padding)

    try:
        excel_bytes = base64.b64decode(file_data)
        wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return {'statusCode': 400, 'headers': {**cors_headers, 'Content-Type': 'application/json'}, 'body': json.dumps({'error': f'Не удалось прочитать Excel: {str(e)}'}), 'isBase64Encoded': False}

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {'statusCode': 400, 'headers': {**cors_headers, 'Content-Type': 'application/json'}, 'body': json.dumps({'error': 'Файл пустой или нет данных'}), 'isBase64Encoded': False}

    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    cur = conn.cursor()

    imported = 0
    created_products = []
    try:
        for row in rows[1:]:
            if not row or not row[0]:
                continue

            name = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            category_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            price = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            description = str(row[3]).strip() if len(row) > 3 and row[3] else ''

            if not name:
                continue

            category = normalize_category(category_raw)

            name_esc = name.replace("'", "''")
            category_esc = category.replace("'", "''")
            price_esc = price.replace("'", "''")
            description_esc = description.replace("'", "''")

            cur.execute(
                f"INSERT INTO products (name, category, price, description, image_url) VALUES ('{name_esc}', '{category_esc}', '{price_esc}', '{description_esc}', '') RETURNING id"
            )
            new_id = cur.fetchone()[0]
            created_products.append({'id': new_id, 'name': name})
            imported += 1

        conn.commit()
    finally:
        cur.close()
        conn.close()

    return {
        'statusCode': 200,
        'headers': {**cors_headers, 'Content-Type': 'application/json'},
        'body': json.dumps({'success': True, 'imported': imported, 'products': created_products}),
        'isBase64Encoded': False
    }